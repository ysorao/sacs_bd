from django.shortcuts import render, redirect
from django.db import connection
from apps.usuario.models import Usuarios
from django.http.response import HttpResponse
from django.views.generic import TemplateView
import openpyxl
from openpyxl import workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from datetime import date
from datetime import datetime, timedelta

class Inicio(TemplateView):
    template_name = 'login.html'

# def GraficasIndex(self, **kwargs):
#     #context=super().get_context_data(**kwargs)
#     #fecha = datetime.strftime(date.today()- timedelta(days=1), '%d/%m/%y')
#     fecha='17/07/2020'
#     cursor=connection.cursor()
#     cursor.execute("exec sp_porcentajeGenBak" + "'"+fecha+"'")
#     pCumplido=cursor.fetchall()
#     print(pCumplido)
#     return render(request, 'index.html', {'Cumplimiento': pCumplido} )


def Consultarusuarios(request):
    cursor=connection.cursor()
    cursor.execute('exec sp_listausuarios')
    resultado_sp=cursor.fetchall()
    return render(request, 'reportes/usuarios.html', {'sp_listausuarios': resultado_sp})

class ReporteUsuarios(TemplateView):
    def get(self,request,*args,**kwargs):
        #usuarios=Usuarios.objects.all()
        cursor=connection.cursor()
        cursor.execute('exec sp_listausuarios')
        usuarios=cursor.fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['B1']='USUARIOS REGISTRADOS'
        ws.merge_cells('B1:D1')
        ws['B2']='NOMBRE'
        ws['C2']='DOCUMENTO'
        ws['D2']='FECHA DE REGISTRO'
        
        for usuario in usuarios:
            ws.append(usuario)
            
        nombre_archivo="Listado_usuarios_registrados.xlsx"
        response= HttpResponse(content_type="application/ms-excel")
        contenido="attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = contenido
        wb.save(response)
        return response

def ConsultarEstadosDB(request):
    cursor=connection.cursor()
    cursor.execute('exec sp_estadosdb')
    listadodb=cursor.fetchall()  
    return render(request, 'reportes/estados_db.html', {'listadodb': listadodb})       

class ReporteEstadosDB(TemplateView):
    def get(self,request,*args,**kwargs):
        cursor=connection.cursor()
        cursor.execute('exec sp_estadosdb')
        estadosdb=cursor.fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active

        #Título de la hoja
        ws.merge_cells('A1:E1')
        ws['A1']= "ESTADO DE LAS BASES DE DATOS"
        
        ws['A1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['A1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['A1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['A1'].font=Font(name='Calibri', size=12, bold=True)

        # Encabezados de tabla       

        ws['A2'].alignment=Alignment(horizontal="center",vertical="center")
        ws['A2'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['A2'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['A2'].font=Font(name='Calibri', size=11, bold=True)
        
        ws['B2'].alignment=Alignment(horizontal="center",vertical="center")
        ws['B2'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['B2'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['B2'].font=Font(name='Calibri', size=11, bold=True)
        
        ws['C2'].alignment=Alignment(horizontal="center",vertical="center")
        ws['C2'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['C2'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['C2'].font=Font(name='Calibri', size=11, bold=True)

        ws['D2'].alignment=Alignment(horizontal="center",vertical="center")
        ws['D2'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['D2'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['D2'].font=Font(name='Calibri', size=11, bold=True)

        ws['E2'].alignment=Alignment(horizontal="center",vertical="center")
        ws['E2'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['E2'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['E2'].font=Font(name='Calibri', size=11, bold=True)

        ws['A2']='SERVIDOR'
        ws['B2']='NOMBRE BD'
        ws['C2']='FECHA DE CREACIÓN'
        ws['D2']='ESTADO'
        ws['E2']='LECURA/ESCRITURA'

        ws.column_dimensions['A'].width=35
        ws.column_dimensions['B'].width=35
        ws.column_dimensions['C'].width=22
        ws.column_dimensions['D'].width=22
        ws.column_dimensions['E'].width=22
        
        #Tabla de datos
        fila=3
        for estado in estadosdb:

            ws.cell(row=fila,column=1).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=1).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=1).font=Font(name='Calibri', size=10)
            ws.cell(row=fila,column=1).value=estado[0]
            
            ws.cell(row=fila,column=2).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=2).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=2).font=Font(name='Calibri', size=10)
            ws.cell(row=fila,column=2).value=estado[1]

            ws.cell(row=fila,column=3).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=3).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=3).font=Font(name='Calibri', size=10)
            ws.cell(row=fila,column=3).value=estado[2]

            ws.cell(row=fila,column=4).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=4).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=4).font=Font(name='Calibri', size=10)
            ws.cell(row=fila,column=4).value=estado[3]

            ws.cell(row=fila,column=5).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=5).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=5).font=Font(name='Calibri', size=10)
            ws.cell(row=fila,column=5).value=estado[4]

            fila+=1

        fechaReporte = datetime.strftime(date.today(), '%d/%m/%y')
        nombre_archivo="Estado actual de las bases de datos-generado el"+' '+fechaReporte+".xlsx"
        response= HttpResponse(content_type="application/ms-excel")
        contenido="attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = contenido
        wb.save(response)
        return response


def ConsultarUltimoBck(request):
    cursor=connection.cursor()
    cursor.execute('exec sp_ultimosbck')
    listadoUltimosBck=cursor.fetchall() 
    return render(request, 'reportes/ultimos_bck.html', {'listadoUBck': listadoUltimosBck})

class ReporteUltimosBck(TemplateView):
    def get(self,request,*args,**kwargs):
        cursor=connection.cursor()
        cursor.execute('exec sp_ultimosbck')
        ultimasCopias=cursor.fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active

        #Título de la hoja
        ws.merge_cells('A1:F1')
        ws['A1']= "ÚLTIMO BACKUP GENERADO A CADA BASE DE DATOS"
        
        ws['A1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['A1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['A1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['A1'].font=Font(name='Calibri', size=12, bold=True)

        # Encabezados de tabla
        ws['A2'].alignment=Alignment(horizontal="center",vertical="center")
        ws['A2'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['A2'].fill=PatternFill(start_color='00FF6600',end_color='00FF6600',fill_type='solid')
        ws['A2'].font=Font(name='Calibri', size=11, bold=True)
        
        ws['B2'].alignment=Alignment(horizontal="center",vertical="center")
        ws['B2'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['B2'].fill=PatternFill(start_color='00FF6600',end_color='00FF6600',fill_type='solid')
        ws['B2'].font=Font(name='Calibri', size=11, bold=True)

        ws['C2'].alignment=Alignment(horizontal="center",vertical="center")
        ws['C2'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['C2'].fill=PatternFill(start_color='00FF6600',end_color='00FF6600',fill_type='solid')
        ws['C2'].font=Font(name='Calibri', size=11, bold=True)
        
        ws['D2'].alignment=Alignment(horizontal="center",vertical="center")
        ws['D2'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['D2'].fill=PatternFill(start_color='00FF6600',end_color='00FF6600',fill_type='solid')
        ws['D2'].font=Font(name='Calibri', size=11, bold=True)

        ws['E2'].alignment=Alignment(horizontal="center",vertical="center")
        ws['E2'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['E2'].fill=PatternFill(start_color='00FF6600',end_color='00FF6600',fill_type='solid')
        ws['E2'].font=Font(name='Calibri', size=11, bold=True)

        ws['F2'].alignment=Alignment(horizontal="center",vertical="center")
        ws['F2'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['F2'].fill=PatternFill(start_color='00FF6600',end_color='00FF6600',fill_type='solid')
        ws['F2'].font=Font(name='Calibri', size=11, bold=True)

        ws['A2']='NOMBRE SERVIDOR'
        ws['B2']='NOMBRE BASE DE DATOS'
        ws['C2']='INICIO BACKUP'
        ws['D2']='FIN BACKUP'
        ws['E2']='TAMAÑO DEL BACKUP (MB)'
        ws['F2']='TIEMPO EJECUCIÓN'

        ws.column_dimensions['A'].width=22
        ws.column_dimensions['B'].width=23
        ws.column_dimensions['C'].width=22
        ws.column_dimensions['D'].width=22
        ws.column_dimensions['E'].width=25
        ws.column_dimensions['F'].width=22
        
        #Tabla de datos
        fila=3
        for resultado in ultimasCopias:

            ws.cell(row=fila,column=1).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=1).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=1).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=1).value=resultado[0]
            
            ws.cell(row=fila,column=2).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=2).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=2).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=2).value=resultado[1]

            ws.cell(row=fila,column=3).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=3).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=3).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=3).value=resultado[2]

            ws.cell(row=fila,column=4).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=4).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=4).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=4).value=resultado[3]
            
            ws.cell(row=fila,column=5).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=5).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=5).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=5).value=resultado[4]

            ws.cell(row=fila,column=6).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=6).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=6).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=6).value=resultado[5]

            fila+=1

        fechaReporte = datetime.strftime(date.today(), '%d/%m/%y')
        nombre_archivo="Último backup por base de datos - generado el"+' '+fechaReporte+".xlsx"
        response= HttpResponse(content_type="application/ms-excel")
        contenido="attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = contenido
        wb.save(response)
        return response

def EjecucionJobBck(request):
    #fecha = request.GET.get('fecha1')
    #if str(request.GET.get('fecha1')) == None:
    #fecha = str(date.today()- timedelta(days=1))
    #else:
        #fecha = request.GET.get('fecha1')
    #return fecha
    #print('el valor del campo es.. '+ str(request.GET.get('fecha1')) )
    fecha = str(date.today()- timedelta(days=1))
    #fecha = '2020-07-17'
    print(fecha)
    cursor=connection.cursor()
    cursor.execute("exec sp_resultadoJobsBck"+ "'"+fecha+"'")
    ejecucionJobs=cursor.fetchall()
    return render(request, 'reportes/jobs_bck.html', {'jobsBck': ejecucionJobs})

class ReporteEjecucionJobBck(TemplateView):
    def get(self,request,*args,**kwargs):
        fecha=request.GET.get('fecha')
        #fecha = '2020-07-17'
        cursor=connection.cursor()
        cursor.execute("exec sp_resultadoJobsBck"+ "'"+fecha+"'")
        ejecuciónJobs=cursor.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active

        # Encabezados de tabla       

        ws['A1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['A1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['A1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['A1'].font=Font(name='Calibri', size=11, bold=True)

        ws['B1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['B1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['B1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['B1'].font=Font(name='Calibri', size=11, bold=True)
        
        ws['C1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['C1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['C1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['C1'].font=Font(name='Calibri', size=11, bold=True)

        ws['D1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['D1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['D1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['D1'].font=Font(name='Calibri', size=11, bold=True)

        ws['E1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['E1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['E1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['E1'].font=Font(name='Calibri', size=11, bold=True)
        
        ws['F1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['F1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['F1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['F1'].font=Font(name='Calibri', size=11, bold=True)

        ws['G1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['G1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['G1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['G1'].font=Font(name='Calibri', size=11, bold=True)

        ws['H1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['H1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['H1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['H1'].font=Font(name='Calibri', size=11, bold=True)

        ws['A1']='RESULTADO'
        ws['B1']='NOMBRE DEL SERVIDOR'
        ws['C1']='FECHA'
        ws['D1']='HORA'
        ws['E1']='NOMBRE DEL JOB'
        ws['F1']='PASO'
        ws['G1']='NOMBRE DEL PASO'
        ws['H1']='MENSAJE'

        ws.column_dimensions['A'].width=12
        ws.column_dimensions['B'].width=30
        ws.column_dimensions['C'].width=12
        ws.column_dimensions['D'].width=12
        ws.column_dimensions['E'].width=32
        ws.column_dimensions['F'].width=10
        ws.column_dimensions['G'].width=43
        ws.column_dimensions['H'].width=45

        #Tabla de datos
        fila=2
        for job in ejecuciónJobs:
            #ws.append(estado)
            ws.cell(row=fila,column=1).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=1).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=1).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=1).value=job[0]
            
            ws.cell(row=fila,column=2).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=2).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=2).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=2).value=job[1]

            ws.cell(row=fila,column=3).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=3).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=3).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=3).value=job[2]

            ws.cell(row=fila,column=4).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=4).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=4).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=4).value=job[3]

            ws.cell(row=fila,column=5).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=5).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=5).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=5).value=job[4]

            ws.cell(row=fila,column=6).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=6).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=6).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=6).value=job[5]

            ws.cell(row=fila,column=7).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=7).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=7).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=7).value=job[6]

            ws.cell(row=fila,column=8).alignment=Alignment(horizontal="left",vertical="center")
            ws.cell(row=fila,column=8).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=8).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=8).value=job[7]

            fila+=1

        fechaReporte = datetime.strftime(date.today(), '%d/%m/%y')
        nombre_archivo="Reporte de ejecución jobs de backup generado el"+' '+fechaReporte+".xlsx"
        response= HttpResponse(content_type="application/ms-excel")
        contenido="attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = contenido
        wb.save(response)
        return response


def GenArchivosBak(request):
    fecha = datetime.strftime(date.today()- timedelta(days=1), '%d/%m/%y')
    #fecha = '17/07/2020'
    cursor=connection.cursor()
    cursor.execute("exec sp_genBak"+ "'"+fecha+"'")
    bakGenerados=cursor.fetchall()
    return render(request, 'reportes/archivos_bak.html', {'bakGen': bakGenerados})

class ReporteArchivosBak(TemplateView):
    def get(self,request,*args,**kwargs):
        fecha=request.GET.get('fecha')
        #fecha = '17/07/2020'
        cursor=connection.cursor()
        cursor.execute("exec sp_genBak"+ "'"+fecha+"'")
        archivosBak=cursor.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active

        # Encabezados de tabla       

        ws['A1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['A1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['A1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['A1'].font=Font(name='Calibri', size=11, bold=True)

        ws['B1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['B1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['B1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['B1'].font=Font(name='Calibri', size=11, bold=True)
        
        ws['C1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['C1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['C1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['C1'].font=Font(name='Calibri', size=11, bold=True)

        ws['D1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['D1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['D1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['D1'].font=Font(name='Calibri', size=11, bold=True)

        ws['E1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['E1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['E1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['E1'].font=Font(name='Calibri', size=11, bold=True)
        
        ws['F1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['F1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['F1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['F1'].font=Font(name='Calibri', size=11, bold=True)


        ws['A1']='NOMBRE SERVIDOR'
        ws['B1']='NOMBRE BASE DE DATOS'
        ws['C1']='FECHA'
        ws['D1']='HORA'
        ws['E1']='TIPO DE BACKUP'
        ws['F1']='RUTA DE ALMACENAMIENTO'

        ws.column_dimensions['A'].width=30
        ws.column_dimensions['B'].width=45
        ws.column_dimensions['C'].width=12
        ws.column_dimensions['D'].width=15
        ws.column_dimensions['E'].width=16
        ws.column_dimensions['F'].width=70

        #Tabla de datos
        fila=2
        for bak in archivosBak:

            ws.cell(row=fila,column=1).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=1).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=1).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=1).value=bak[1]
            
            ws.cell(row=fila,column=2).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=2).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=2).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=2).value=bak[2]

            ws.cell(row=fila,column=3).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=3).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=3).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=3).value=bak[3]

            ws.cell(row=fila,column=4).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=4).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=4).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=4).value=bak[4]

            ws.cell(row=fila,column=5).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=5).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=5).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=5).value=bak[5].strip()

            ws.cell(row=fila,column=6).alignment=Alignment(horizontal="left",vertical="center")
            ws.cell(row=fila,column=6).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=6).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=6).value=bak[6].strip()

            fila+=1

        fechaReporte = datetime.strftime(date.today(), '%d/%m/%y')
        nombre_archivo="Archivos .bak generados el"+' '+fechaReporte+".xlsx"
        response= HttpResponse(content_type="application/ms-excel")
        contenido="attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = contenido
        wb.save(response)
        return response

def CumplimientoBck(request):
    fecha = datetime.strftime(date.today()- timedelta(days=1), '%d/%m/%y')
    #fecha = '17/07/2020'
    cursor=connection.cursor()
    cursor.execute("exec sp_historicoBck"+ "'"+fecha+"'")
    
    resultados=cursor.fetchall()
    return render(request, 'reportes/cumplimiento_bck.html', {'resultadosCump': resultados})


class ReporteCumplimientoBck(TemplateView):
    def get(self,request,*args,**kwargs):
        fecha=request.GET.get('fecha')
        #fecha = '17/07/2020'
        cursor=connection.cursor()
        cursor.execute("exec sp_historicoBck"+ "'"+fecha+"'")
        cumplimiento=cursor.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active

        # Encabezados de tabla       

        ws['A1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['A1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['A1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['A1'].font=Font(name='Calibri', size=11, bold=True)

        ws['B1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['B1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['B1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['B1'].font=Font(name='Calibri', size=11, bold=True)
        
        ws['C1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['C1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['C1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['C1'].font=Font(name='Calibri', size=11, bold=True)

        ws['D1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['D1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['D1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['D1'].font=Font(name='Calibri', size=11, bold=True)

        ws['E1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['E1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['E1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['E1'].font=Font(name='Calibri', size=11, bold=True)
        
        ws['F1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['F1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['F1'].fill=PatternFill(start_color='003366FF',end_color='003366FF',fill_type='solid')
        ws['F1'].font=Font(name='Calibri', size=11, bold=True)


        ws['A1']='Fecha'
        ws['B1']='IP Servidor'
        ws['C1']='Nombre Servidor'
        ws['D1']='Copias ejecutadas'
        ws['E1']='Copias Esperadas'
        ws['F1']='Mensaje'

        ws.column_dimensions['A'].width=22
        ws.column_dimensions['B'].width=35
        ws.column_dimensions['C'].width=45
        ws.column_dimensions['D'].width=20
        ws.column_dimensions['E'].width=20
        ws.column_dimensions['F'].width=20

        #Tabla de datos
        fila=2
        for c in cumplimiento:

            ws.cell(row=fila,column=1).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=1).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=1).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=1).value=c[0]
            
            ws.cell(row=fila,column=2).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=2).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=2).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=2).value=c[1]

            ws.cell(row=fila,column=3).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=3).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=3).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=3).value=c[2]

            ws.cell(row=fila,column=4).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=4).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=4).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=4).value=c[3]

            ws.cell(row=fila,column=5).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=5).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=5).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=5).value=c[4]

            ws.cell(row=fila,column=6).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(row=fila,column=6).border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),
                                top=Side(border_style="thin"),bottom=Side(border_style="thin"))
            ws.cell(row=fila,column=6).font=Font(name='Calibri', size=10, bold=True)
            ws.cell(row=fila,column=6).value=c[5].strip()

            fila+=1

        fechaReporte = datetime.strftime(date.today(), '%d/%m/%y')
        nombre_archivo="Cumplimiento de copias por instancia para el"+' '+fecha+' '+'generado el'+' '+fechaReporte+".xlsx"
        response= HttpResponse(content_type="application/ms-excel")
        contenido="attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = contenido
        wb.save(response)
        return response
